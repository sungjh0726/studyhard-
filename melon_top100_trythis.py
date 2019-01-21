import openpyxl
import csv, codecs
from bs4 import BeautifulSoup
import requests
import json
from time import sleep
from pprint import pprint
from openpyxl.chart import (
    Reference, BarChart, Series, ScatterChart
)
import os
import os.path as path
import urllib.parse as parse
from PIL import Image

# 1) 지난 시간에 작성한 meltop100.csv 파일을 읽어, meltop100.xlsx 로 저장하시오.
#  (단, 랭킹, 좋아요, 좋아요차이 컬럼은 숫자형식으로 저장 할 것!)


fp = codecs.open("./melon_top_100.csv", "r", encoding = "utf-8")

reader = csv.reader(fp, delimiter=',', quotechar='"')

book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = "첫번째 시트"

# for j in range (0,5):
#     for i, cells in enumerate(reader):
#         sheet.cell(row = (i+1), column = (j+1)) = cells[j]


for i, cells in enumerate(reader):
    sheet1.cell(row= (i+1), column = 1).value = cells[0]
    sheet1.cell(row= (i+1), column = 2).value = cells[1]
    sheet1.cell(row= (i+1), column = 3).value = cells[2]
    sheet1.cell(row= (i+1), column = 4).value = cells[3]
    sheet1.cell(row= (i+1), column = 5).value = cells[4]


#------------------------------------------------------------------------------------

# 2)멜론 Top100 곡들의 `앨범 재킷파일`을 다운받아,meltop100.xlsx 파일의 두번째 시트에 랭킹순으로 작성하시오. (단, 이미지파일의 크기는 축소하여 보기 좋게 작성 할 것!)


url = "https://www.melon.com/chart/index.htm"
headers = {
    'Referer': 'https://www.melon.com/',
    'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36'
}
html = requests.get(url, headers = headers).text
soup = BeautifulSoup(html, 'html.parser')
song_lsts = soup.select('div#tb_list table tbody tr[data-song-no]')

# sheet 만들기

load_book = openpyxl.load_workbook("./meltop100.xlsx")
sheet2 = load_book.create_sheet()
sheet2.title = "Sheet 2"

# 이미지 다운받아서 xls에 추가하기
n = 1
for song in song_lsts:
    src = song.select_one('td:nth-of-type(4) div.wrap a img[src]').text
    rank = song.select_one('td:nth-of-type(2) div.wrap.t_center span.rank').text
    src = song.select_one('td:nth-of-type(4) div.wrap a img[src]').attrs['src']
    saveFile = "./images/rank{}.png".format(rank)
    ur.urlretrieve(src, saveFile)
    img = Image.open(saveFile)
    resized_img = img.resize((139, 139))
    resized_img.save('new_rank{}.png'.format(rank))
    final_img = openpyxl.drawing.image.Image('new_rank{}.png'.format(rank))
    if n == 1:
        sheet2.add_image(final_img, 'A{}'.format(n))
        load_book.save("./meltop100.xlsx")
    else:
        sheet2.add_image(final_img, 'A{}'.format(n))
        load_book.save("./meltop100.xlsx")
    n = n + 7
#--------------------------------------------------------------------------------------------
    # 3) 상위 Top10의 `좋아요 수`는 BarChart로, `좋아요 차이 수`는 ScatterChart로 세번째 시트에 작성하시오.
 
sheet3 = book.create_sheet()
sheet3.title = "세번째 시트"
        

datax = Reference(sheet1, min_col=4,               
		min_row=2, max_col=4, max_row=11)
categs = Reference(sheet1, min_col=2, min_row=2, max_row=11)

chart = BarChart()
chart.add_data(data=datax)
chart.set_categories(categs)

chart.legend = None  # 범례
chart.varyColors = True
chart.title = "좋아요 차트"

sheet3.add_chart(chart, "A3")
